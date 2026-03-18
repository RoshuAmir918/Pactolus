#!/usr/bin/env python3
import csv
import json
import math
import os
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


CLAIMS_HINTS = {"claim", "incurred", "paid", "loss", "accident", "reported"}
POLICIES_HINTS = {"policy", "premium", "exposure", "effective", "expiration", "insured"}
TRIANGLE_HINTS = {"triangle", "development", "dev", "accident year", "ay", "dy", "age"}
CONTRACT_HINTS = {"treaty", "retention", "attachment", "cession", "reinsurer", "cedant"}
SUMMARY_SHEET_HINTS = {
    "summary",
    "totals",
    "total",
    "pivot",
    "dashboard",
    "analysis",
    "overview",
    "recap",
}
META_SHEET_HINTS = {
    "meta",
    "metadata",
    "readme",
    "notes",
    "lookup",
    "mapping",
    "dictionary",
    "assumption",
    "glossary",
}


def parse_positive_int_env(name: str, fallback: int) -> int:
    raw = os.getenv(name)
    if raw is None:
        return fallback
    try:
        parsed = int(raw)
    except Exception:
        return fallback
    return parsed if parsed > 0 else fallback


def normalize_text(value: Any) -> str:
    return str(value or "").strip().lower()


def is_number(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return not isinstance(value, bool) and not math.isnan(float(value))
    text = normalize_text(value).replace(",", "").replace("$", "").replace("%", "")
    if not text:
        return False
    try:
        float(text)
        return True
    except Exception:
        return False


def to_float(value: Any) -> Optional[float]:
    if not is_number(value):
        return None
    text = normalize_text(value).replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(text)
    except Exception:
        return None


def keyword_score(headers: List[str], hints: set[str]) -> int:
    header_blob = " ".join([normalize_text(h) for h in headers])
    return sum(1 for hint in hints if hint in header_blob)


def build_segment_manifest(headers: List[str], rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not rows or not headers:
        return []

    numeric_columns = [h for h in headers if sum(1 for r in rows[:2000] if is_number(r.get(h))) >= 20]
    candidates = []
    for header in headers:
        values = [str(r.get(header, "")).strip() for r in rows if str(r.get(header, "")).strip()]
        if not values:
            continue
        unique_values = set(values)
        if 1 < len(unique_values) <= 100:
            candidates.append(header)

    manifests: List[Dict[str, Any]] = []
    for dimension in candidates[:10]:
        buckets: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            value = str(row.get(dimension, "")).strip()
            if not value:
                continue
            if value not in buckets:
                buckets[value] = {"value": value, "rowCount": 0, "numericTotals": defaultdict(float)}
            bucket = buckets[value]
            bucket["rowCount"] += 1
            for col in numeric_columns[:8]:
                num = to_float(row.get(col))
                if num is not None:
                    bucket["numericTotals"][col] += num

        manifests.append(
            {
                "dimension": dimension,
                "values": [
                    {
                        "value": bucket["value"],
                        "rowCount": bucket["rowCount"],
                        "numericTotals": dict(bucket["numericTotals"]),
                    }
                    for bucket in sorted(
                        buckets.values(), key=lambda item: item["rowCount"], reverse=True
                    )[:100]
                ],
            }
        )
    return manifests


def build_aggregate_stats(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, Any]:
    numeric_cols = [h for h in headers if sum(1 for r in rows[:2000] if is_number(r.get(h))) >= 20]
    numeric_stats: Dict[str, Any] = {}
    for col in numeric_cols[:20]:
        values = [to_float(r.get(col)) for r in rows]
        values = [v for v in values if v is not None]
        if not values:
            continue
        values.sort()
        numeric_stats[col] = {
            "count": len(values),
            "min": values[0],
            "max": values[-1],
            "mean": sum(values) / len(values),
            "median": statistics.median(values),
            "p90": values[min(len(values) - 1, int(len(values) * 0.9))],
            "p95": values[min(len(values) - 1, int(len(values) * 0.95))],
            "p99": values[min(len(values) - 1, int(len(values) * 0.99))],
        }
    return {"rowCount": len(rows), "columnCount": len(headers), "numericColumns": numeric_stats}


def build_quality_flags(headers: List[str], rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    flags: List[Dict[str, Any]] = []
    lower_headers = [normalize_text(h) for h in headers]
    critical = [h for h in headers if any(k in normalize_text(h) for k in ["claim", "policy", "accident", "date"])]
    for col in critical[:8]:
        missing = sum(1 for r in rows if not str(r.get(col, "")).strip())
        rate = (missing / max(1, len(rows))) * 100
        if rate > 1:
            flags.append(
                {
                    "code": "MISSING_CRITICAL_VALUES",
                    "column": col,
                    "severity": "high" if rate > 10 else "medium",
                    "details": {"missingCount": missing, "missingPercent": round(rate, 2)},
                }
            )

    id_cols = [h for h in headers if any(k in normalize_text(h) for k in ["claim id", "claim number", "policy number"])]
    for col in id_cols[:2]:
        values = [str(r.get(col, "")).strip() for r in rows if str(r.get(col, "")).strip()]
        counts = Counter(values)
        duplicate_count = sum(c - 1 for c in counts.values() if c > 1)
        if duplicate_count > 0:
            flags.append(
                {
                    "code": "DUPLICATE_IDS",
                    "column": col,
                    "severity": "medium",
                    "details": {"duplicateCount": duplicate_count},
                }
            )

    non_negative = [h for h in headers if any(k in normalize_text(h) for k in ["premium", "paid", "incurred", "loss", "exposure"])]
    for col in non_negative[:10]:
        negatives = sum(1 for r in rows if (to_float(r.get(col)) or 0) < 0)
        if negatives > 0:
            flags.append(
                {
                    "code": "NEGATIVE_VALUES",
                    "column": col,
                    "severity": "medium",
                    "details": {"negativeCount": negatives},
                }
            )

    if not lower_headers:
        flags.append({"code": "NO_HEADERS", "severity": "high", "details": {}})
    return flags


def extract_triangle_from_table(
    headers: List[str], rows: List[Dict[str, Any]], sheet_index: int, sheet_name: str
) -> Optional[Dict[str, Any]]:
    if len(headers) < 3 or len(rows) < 3:
        return None

    first = headers[0]
    dev_headers = headers[1:30]
    matrix = []
    numeric_seen = 0
    for row in rows[:200]:
        label = str(row.get(first, "")).strip()
        values = []
        has_num = False
        for h in dev_headers:
            num = to_float(row.get(h))
            values.append(num)
            if num is not None:
                has_num = True
                numeric_seen += 1
        if label or has_num:
            matrix.append({"label": label, "values": values})

    if numeric_seen < 20:
        return None

    triangle_type = "unknown"
    header_blob = " ".join(normalize_text(h) for h in headers)
    if "paid" in header_blob:
        triangle_type = "paid"
    elif "incurred" in header_blob:
        triangle_type = "incurred"
    elif "reported" in header_blob:
        triangle_type = "reported"
    elif "ultimate" in header_blob:
        triangle_type = "ultimate"

    return {
        "sheetIndex": sheet_index,
        "title": f"{sheet_name} triangle",
        "segmentLabel": None,
        "triangleType": triangle_type,
        "rowStart": 1,
        "rowEnd": len(matrix) + 1,
        "colStart": 1,
        "colEnd": len(dev_headers) + 1,
        "headerLabelsJson": {"rowLabelHeader": first, "developmentHeaders": dev_headers},
        "normalizedTriangleJson": {"rows": matrix, "rowLabelHeader": first, "developmentHeaders": dev_headers},
        "confidence": 0.75,
    }


def classify_table(headers: List[str], row_count: int) -> Tuple[str, str, int, int, int, int]:
    c = keyword_score(headers, CLAIMS_HINTS)
    p = keyword_score(headers, POLICIES_HINTS)
    t = keyword_score(headers, TRIANGLE_HINTS)
    x = keyword_score(headers, CONTRACT_HINTS)

    sheet_type = "unknown"
    # Avoid misclassifying large claims bordereaux sheets as triangles.
    # Triangle tabs are usually compact and development-heavy, while claims tabs
    # often contain one or two overlap keywords like "accident"/"reported".
    if t >= 3 and row_count <= 250 and c <= 2 and p <= 2:
        sheet_type = "triangle_like"
    elif c >= p and c >= 2:
        sheet_type = "claims_like"
    elif p > c and p >= 2:
        sheet_type = "policies_like"
    elif x >= 2:
        sheet_type = "other"
    elif row_count > 200:
        sheet_type = "claims_like"
    else:
        sheet_type = "other"

    return sheet_type, sheet_type, c, p, t, x


def has_any_hint(text: str, hints: set[str]) -> bool:
    return any(hint in text for hint in hints)


def find_headers(headers: List[str], keywords: List[str]) -> List[str]:
    matches = []
    for header in headers:
        lowered = normalize_text(header)
        if any(keyword in lowered for keyword in keywords):
            matches.append(header)
    return matches


def best_unique_ratio(rows: List[Dict[str, Any]], headers: List[str]) -> float:
    if not rows or not headers:
        return 0.0
    best = 0.0
    sample_rows = rows[:2000]
    for header in headers:
        values = [str(r.get(header, "")).strip().lower() for r in sample_rows if str(r.get(header, "")).strip()]
        if len(values) < 20:
            continue
        ratio = len(set(values)) / max(1, len(values))
        if ratio > best:
            best = ratio
    return best


def average_numeric_rate(rows: List[Dict[str, Any]], headers: List[str]) -> float:
    if not rows or not headers:
        return 0.0
    sample_rows = rows[:2000]
    rates = []
    for header in headers[:8]:
        non_empty = [r.get(header) for r in sample_rows if str(r.get(header, "")).strip()]
        if len(non_empty) < 20:
            continue
        numeric = sum(1 for value in non_empty if is_number(value))
        rates.append(numeric / max(1, len(non_empty)))
    if not rates:
        return 0.0
    return sum(rates) / len(rates)


def derive_sheet_role(
    sheet_type: str,
    profile_role: str,
    sheet_name: str,
    row_count_estimate: int,
    triangle_score: int,
) -> str:
    normalized_name = normalize_text(sheet_name)
    if sheet_type == "triangle_like":
        if (
            triangle_score >= 4
            or row_count_estimate >= 40
            or "triangle" in normalized_name
            or "development" in normalized_name
            or "loss dev" in normalized_name
        ):
            return "triangle_data"
        return "summary_triangle"
    if profile_role == "transactional":
        return "transactional_data"
    if profile_role == "summary_or_meta":
        return "summary_or_meta"
    return "other"


def build_sheet_about_payload(
    sheet_index: int,
    sheet_name: str,
    sheet_type: str,
    row_count_estimate: int,
    column_count: int,
    profile: Dict[str, Any],
    c_score: int,
    p_score: int,
    t_score: int,
    x_score: int,
) -> Dict[str, Any]:
    role = derive_sheet_role(
        sheet_type=sheet_type,
        profile_role=str(profile.get("role") or "other"),
        sheet_name=sheet_name,
        row_count_estimate=row_count_estimate,
        triangle_score=t_score,
    )
    return {
        "version": 1,
        "sheetIndex": sheet_index,
        "sheetName": sheet_name,
        "role": role,
        "score": int(profile.get("score") or 0),
        "confidence": float(profile.get("confidence") or 0.0),
        "reasons": profile.get("reasons") or [],
        "rowCountEstimate": row_count_estimate,
        "columnCount": column_count,
        "selectedAsPrimary": False,
        "selectedForExtraction": role in {"triangle_data", "summary_triangle"},
        "signals": {
            "claimsScore": c_score,
            "policiesScore": p_score,
            "triangleScore": t_score,
            "contractScore": x_score,
        },
    }


def score_record_level_sheet(
    sheet_name: str, headers: List[str], rows: List[Dict[str, Any]]
) -> Dict[str, Any]:
    row_count = len(rows)
    normalized_sheet_name = normalize_text(sheet_name)
    header_blob = " ".join(normalize_text(h) for h in headers)

    claim_id_headers = find_headers(headers, ["claim id", "claim number", "claim no", "claim #"])
    policy_id_headers = find_headers(headers, ["policy number", "policy no", "policy #", "policy id"])
    date_headers = find_headers(
        headers,
        [
            "accident date",
            "loss date",
            "reported date",
            "report date",
            "closed date",
            "open date",
            "effective",
            "expiration",
            "inception",
        ],
    )
    amount_headers = find_headers(
        headers,
        ["paid", "incurred", "reserve", "loss", "premium", "exposure", "amount"],
    )
    status_headers = find_headers(headers, ["status", "open", "closed"])

    unique_ratio = best_unique_ratio(rows, claim_id_headers + policy_id_headers)
    numeric_rate = average_numeric_rate(rows, amount_headers)

    score = 0
    reasons: List[str] = []

    if row_count >= 300:
        score += 4
        reasons.append("large_row_count")
    elif row_count >= 120:
        score += 2
        reasons.append("moderate_row_count")
    elif row_count < 40:
        score -= 3
        reasons.append("low_row_count")

    if claim_id_headers:
        score += 3
        reasons.append("has_claim_id_headers")
    if policy_id_headers:
        score += 2
        reasons.append("has_policy_id_headers")
    if date_headers:
        score += 2
        reasons.append("has_date_headers")
    if len(amount_headers) >= 2:
        score += 2
        reasons.append("has_financial_headers")
    elif len(amount_headers) == 1:
        score += 1
        reasons.append("has_single_financial_header")
    if status_headers:
        score += 1
        reasons.append("has_status_headers")

    if unique_ratio >= 0.7:
        score += 2
        reasons.append("high_identifier_cardinality")
    elif row_count >= 120 and unique_ratio > 0 and unique_ratio < 0.2:
        score -= 2
        reasons.append("low_identifier_cardinality")

    if numeric_rate >= 0.75:
        score += 1
        reasons.append("high_numeric_consistency")

    summary_name_hits = sum(1 for hint in SUMMARY_SHEET_HINTS if hint in normalized_sheet_name)
    summary_header_hits = sum(
        1 for hint in {"total", "summary", "grand total", "subtotal", "pivot"} if hint in header_blob
    )
    summary_like = summary_name_hits >= 1 or summary_header_hits >= 2 or (
        row_count < 250 and summary_header_hits >= 1
    )
    meta_like = has_any_hint(normalized_sheet_name, META_SHEET_HINTS)

    if summary_like:
        score -= 4
        reasons.append("summary_like_sheet")
    if meta_like:
        score -= 4
        reasons.append("metadata_like_sheet")

    role = "other"
    if score >= 8 and row_count >= 300 and not meta_like:
        # Large row-level sheets with strong transactional signals should win
        # even if they contain some summary-like wording in headers.
        role = "transactional"
    elif score >= 5 and not summary_like and not meta_like:
        role = "transactional"
    elif summary_like or meta_like:
        role = "summary_or_meta"

    confidence = max(0.0, min(1.0, (score + 6) / 14))

    return {
        "sheetName": sheet_name,
        "rowCount": row_count,
        "score": score,
        "confidence": round(confidence, 4),
        "role": role,
        "reasons": reasons[:8],
        "claimIdHeaders": claim_id_headers[:6],
        "policyIdHeaders": policy_id_headers[:6],
        "dateHeaders": date_headers[:8],
        "amountHeaders": amount_headers[:10],
    }


def read_csv_table(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        rows = []
        for row in reader:
            rows.append({h: row.get(h) for h in headers})
            if len(rows) >= 50000:
                break

    sheet_type, ai_class, c, p, t, x = classify_table(headers, len(rows))
    sheet_profile = score_record_level_sheet("main", headers, rows)
    sheet_about = build_sheet_about_payload(
        sheet_index=0,
        sheet_name="main",
        sheet_type=sheet_type,
        row_count_estimate=len(rows),
        column_count=len(headers),
        profile=sheet_profile,
        c_score=c,
        p_score=p,
        t_score=t,
        x_score=x,
    )

    sheet = {
        "sheetName": "main",
        "sheetIndex": 0,
        "sheetType": sheet_type,
        "aiClassification": ai_class,
        "aiConfidence": 0.85 if sheet_type != "other" else 0.55,
        "usedRangeJson": {"startRow": 1, "startCol": 1, "endRow": len(rows) + 1, "endCol": len(headers)},
        "headersJson": headers,
        "sampleRowsJson": rows[:500],
        "rowCountEstimate": len(rows),
        "detectedTablesJson": [
            {
                "title": "main",
                "bounds": {"startRow": 1, "startCol": 1, "endRow": len(rows) + 1, "endCol": len(headers)},
                "confidence": 0.9,
                "possibleType": sheet_type,
            }
        ],
        "sheetAboutJson": sheet_about,
        "searchText": " ".join((headers[:20] + ["main"]))[:10000],
    }

    route = "deterministic_claims_policies"
    if sheet_type == "triangle_like":
        route = "hybrid_triangles"
    elif x >= 2 and len(rows) < 200:
        route = "claude_contract"

    triangles = []
    if route == "hybrid_triangles":
        tri = extract_triangle_from_table(headers, rows, 0, "main")
        if tri:
            triangles.append(tri)
    sheet_about["selectedAsPrimary"] = route != "hybrid_triangles"
    sheet_about["selectedForExtraction"] = route == "hybrid_triangles"

    doc_type = "other"
    if route == "hybrid_triangles":
        doc_type = "loss_triangles"
    elif sheet_type == "claims_like":
        doc_type = "claims"
    elif sheet_type == "policies_like":
        doc_type = "policies"

    aggregate_stats = build_aggregate_stats(headers, rows)
    aggregate_stats["sourceSheet"] = {
        "sheetIndex": 0,
        "sheetName": "main",
        "selectionMethod": "single_sheet_csv",
    }

    result = {
        "route": route,
        "document": {
            "documentType": doc_type,
            "aiClassification": doc_type if doc_type != "other" else "unknown",
            "aiConfidence": 0.8 if doc_type != "other" else 0.5,
            "searchText": " ".join(headers[:40])[:10000],
        },
        "sheets": [sheet],
        "triangles": triangles,
        "deterministic": {
            "segmentManifest": build_segment_manifest(headers, rows) if route == "deterministic_claims_policies" else [],
            "aggregateStats": aggregate_stats,
            "qualityFlags": build_quality_flags(headers, rows),
        },
    }
    return result


def read_xlsx_tables(path: Path) -> Dict[str, Any]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return {
            "route": "claude_contract",
            "document": {
                "documentType": "other",
                "aiClassification": "unknown",
                "aiConfidence": 0.2,
                "searchText": "openpyxl not installed; fallback to Claude-first",
            },
            "sheets": [],
            "triangles": [],
            "deterministic": {"segmentManifest": [], "aggregateStats": {}, "qualityFlags": []},
            "notes": ["openpyxl_missing"],
        }

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    row_scan_limit = parse_positive_int_env("INGESTION_XLSX_ROW_SCAN_LIMIT", 5000)
    extracted_sheets: List[Dict[str, Any]] = []
    score_totals = {"claims_like": 0, "policies_like": 0, "triangle_like": 0, "other": 0}
    all_rows_for_deterministic: List[Dict[str, Any]] = []
    canonical_headers: List[str] = []
    deterministic_sheet_index: Optional[int] = None
    deterministic_sheet_name: Optional[str] = None
    deterministic_sheet_score: Optional[int] = None
    fallback_rows: List[Dict[str, Any]] = []
    fallback_headers: List[str] = []
    fallback_sheet_index: Optional[int] = None
    fallback_sheet_name: Optional[str] = None
    sheet_profiles: List[Dict[str, Any]] = []
    triangles = []
    text_like_count = 0

    for idx, ws in enumerate(wb.worksheets):
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([cell for cell in row])
            if len(rows) >= row_scan_limit:
                break

        if not rows:
            continue

        header_idx = 0
        best_score = -1
        for i, row in enumerate(rows[:20]):
            non_empty = [c for c in row if str(c or "").strip()]
            score = len(non_empty)
            if score > best_score:
                best_score = score
                header_idx = i

        headers = [str(c).strip() if str(c or "").strip() else f"col_{j+1}" for j, c in enumerate(rows[header_idx])]
        table_rows: List[Dict[str, Any]] = []
        for data_row in rows[header_idx + 1 :]:
            # Keep sparse title/section rows (single non-empty cell) because they often
            # carry triangle-type cues like "Paid"/"Incurred"/"Reported".
            if sum(1 for c in data_row if str(c or "").strip()) <= 0:
                continue
            item = {headers[j]: data_row[j] if j < len(data_row) else None for j in range(len(headers))}
            table_rows.append(item)

        # ws.max_row is the workbook's native row count and can exceed the in-memory
        # scan window; use it for a better row-count estimate.
        max_row_value = ws.max_row if isinstance(ws.max_row, int) else 0
        row_count_estimate = max(
            len(table_rows),
            max(0, max_row_value - (header_idx + 1)),
        )

        sheet_type, ai_class, c, p, t, x = classify_table(headers, len(table_rows))
        if x >= 3 and len(table_rows) < 100:
            text_like_count += 1
        score_totals[sheet_type if sheet_type in score_totals else "other"] += 1

        if sheet_type in {"claims_like", "policies_like"}:
            if len(table_rows) > len(fallback_rows):
                fallback_rows = table_rows
                fallback_headers = headers
                fallback_sheet_index = idx
                fallback_sheet_name = ws.title

        profile = score_record_level_sheet(ws.title, headers, table_rows)
        profile["sheetIndex"] = idx
        sheet_profiles.append(profile)
        if profile["role"] == "transactional":
            better_candidate = (
                deterministic_sheet_score is None
                or profile["score"] > deterministic_sheet_score
                or (
                    profile["score"] == deterministic_sheet_score
                    and len(table_rows) > len(all_rows_for_deterministic)
                )
            )
            if better_candidate:
                all_rows_for_deterministic = table_rows
                canonical_headers = headers
                deterministic_sheet_index = idx
                deterministic_sheet_name = ws.title
                deterministic_sheet_score = profile["score"]

        if sheet_type == "triangle_like":
            tri = extract_triangle_from_table(headers, table_rows, idx, ws.title)
            if tri:
                triangles.append(tri)

        extracted_sheets.append(
            {
                "sheetName": ws.title,
                "sheetIndex": idx,
                "sheetType": sheet_type,
                "aiClassification": ai_class,
                "aiConfidence": 0.85 if sheet_type != "other" else 0.55,
                "usedRangeJson": {"startRow": 1, "startCol": 1, "endRow": len(rows), "endCol": len(headers)},
                "headersJson": headers,
                "sampleRowsJson": table_rows[:500],
                "rowCountEstimate": row_count_estimate,
                "detectedTablesJson": [
                    {
                        "title": ws.title,
                        "bounds": {"startRow": header_idx + 1, "startCol": 1, "endRow": len(rows), "endCol": len(headers)},
                        "confidence": 0.8,
                        "possibleType": sheet_type,
                    }
                ],
                "sheetAboutJson": build_sheet_about_payload(
                    sheet_index=idx,
                    sheet_name=ws.title,
                    sheet_type=sheet_type,
                    row_count_estimate=row_count_estimate,
                    column_count=len(headers),
                    profile=profile,
                    c_score=c,
                    p_score=p,
                    t_score=t,
                    x_score=x,
                ),
                "searchText": " ".join([ws.title] + headers[:20])[:10000],
            }
        )

    route = "deterministic_claims_policies"
    if triangles:
        route = "hybrid_triangles"
    elif text_like_count >= max(1, len(extracted_sheets) // 2):
        route = "claude_contract"

    if not canonical_headers and fallback_headers:
        canonical_headers = fallback_headers
        all_rows_for_deterministic = fallback_rows
        deterministic_sheet_index = fallback_sheet_index
        deterministic_sheet_name = fallback_sheet_name

    for sheet in extracted_sheets:
        current_index = sheet.get("sheetIndex")
        sheet_about = sheet.get("sheetAboutJson")
        if not isinstance(sheet_about, dict):
            continue
        sheet_about["selectedAsPrimary"] = (
            route == "deterministic_claims_policies"
            and deterministic_sheet_index is not None
            and current_index == deterministic_sheet_index
        )
        sheet_about["selectedForExtraction"] = route == "hybrid_triangles" and (
            sheet.get("sheetType") == "triangle_like"
        )

    doc_type = "other"
    if route == "hybrid_triangles":
        doc_type = "loss_triangles"
    elif score_totals["claims_like"] >= score_totals["policies_like"] and score_totals["claims_like"] > 0:
        doc_type = "claims"
    elif score_totals["policies_like"] > 0:
        doc_type = "policies"

    aggregate_stats = (
        build_aggregate_stats(canonical_headers, all_rows_for_deterministic) if canonical_headers else {}
    )
    aggregate_stats["sourceSheet"] = {
        "sheetIndex": deterministic_sheet_index,
        "sheetName": deterministic_sheet_name,
        "selectionMethod": "deterministic_scoring",
        "selectionScore": deterministic_sheet_score,
        "candidateCount": len(sheet_profiles),
    }
    aggregate_stats["sheetSelectionDiagnostics"] = sorted(
        [
            {
                "sheetIndex": profile.get("sheetIndex"),
                "sheetName": profile.get("sheetName"),
                "role": profile.get("role"),
                "score": profile.get("score"),
                "confidence": profile.get("confidence"),
                "rowCount": profile.get("rowCount"),
                "reasons": profile.get("reasons"),
            }
            for profile in sheet_profiles
        ],
        key=lambda item: (
            int(item.get("score") or -9999),
            int(item.get("rowCount") or 0),
        ),
        reverse=True,
    )[:10]

    return {
        "route": route,
        "document": {
            "documentType": doc_type,
            "aiClassification": doc_type if doc_type != "other" else "unknown",
            "aiConfidence": 0.8 if doc_type != "other" else 0.5,
            "searchText": " ".join(s["searchText"] for s in extracted_sheets[:10])[:10000],
        },
        "sheets": extracted_sheets,
        "triangles": triangles,
        "deterministic": {
            "segmentManifest": build_segment_manifest(canonical_headers, all_rows_for_deterministic)
            if route == "deterministic_claims_policies"
            else [],
            "aggregateStats": aggregate_stats,
            "qualityFlags": build_quality_flags(canonical_headers, all_rows_for_deterministic)
            if canonical_headers
            else [],
        },
    }


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "missing input json path"}))
        return 1
    input_path = Path(sys.argv[1])
    data = json.loads(input_path.read_text(encoding="utf-8"))
    file_path = Path(data["filePath"])
    ext = normalize_text(data.get("fileExtension"))

    if ext == "csv":
        out = read_csv_table(file_path)
    elif ext == "xlsx":
        out = read_xlsx_tables(file_path)
    else:
        out = {
            "route": "claude_contract",
            "document": {
                "documentType": "other",
                "aiClassification": "unknown",
                "aiConfidence": 0.4,
                "searchText": f"unsupported deterministic type: {ext}",
            },
            "sheets": [],
            "triangles": [],
            "deterministic": {"segmentManifest": [], "aggregateStats": {}, "qualityFlags": []},
        }

    print(json.dumps(out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
